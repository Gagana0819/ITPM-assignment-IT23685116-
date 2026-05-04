"""
Populates 'Assignment 1 - Test cases.xlsx' with 50 Singlish->Sinhala NEGATIVE test cases
covering edge cases, invalid inputs, stress inputs, and boundary conditions.
All 50 cases are negative test cases (Neg_0001 – Neg_0050) with Short/Medium/Long variety.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

EXCEL_PATH = r"D:\3Y1S\ITPM\ITPM-assignment-IT23685116-\Assignment 1 - Test cases.xlsx"

# 50 NEGATIVE test cases:
# (tc_id, length_type, singlish_input, expected_sinhala, type_covered, rationale)
TEST_CASES = [

    # ── 1. EMPTY / WHITESPACE (4) ──────────────────────────────────────────────
    ("Neg_0001", "Short",
     "",
     "",
     "Empty / Whitespace",
     "Completely empty string – tool should return empty or show an error gracefully without crashing."),

    ("Neg_0002", "Short",
     " ",
     "",
     "Empty / Whitespace",
     "Single space character – whitespace-only input should be treated as empty and handled gracefully."),

    ("Neg_0003", "Short",
     "     ",
     "",
     "Empty / Whitespace",
     "Multiple spaces only – tool must not produce any transliteration output for blank-equivalent input."),

    ("Neg_0004", "Short",
     "\t\t",
     "",
     "Empty / Whitespace",
     "Tab characters only – invisible control characters should not trigger a translation result."),

    # ── 2. SPECIAL CHARACTERS / SYMBOL OVERLOAD (6) ────────────────────────────
    ("Neg_0005", "Short",
     "!@#$%^&*()",
     "",
     "Special Characters",
     "Pure ASCII special-character symbols – no Singlish words present; output should be empty or unchanged."),

    ("Neg_0006", "Short",
     "???!!!",
     "",
     "Special Characters",
     "Repeated punctuation marks only – no translatable content; tool should handle gracefully."),

    ("Neg_0007", "Medium",
     "kohomada??? how are you!!! ##ok##",
     "කොහොමද??? how are you!!! ##ok##",
     "Special Characters",
     "Excessive punctuation mixed with valid Singlish – tests whether repeated symbols distort transliteration."),

    ("Neg_0008", "Short",
     "---===///",
     "",
     "Special Characters",
     "Non-alphanumeric symbol sequence – should not crash the translator; expects pass-through or empty."),

    ("Neg_0009", "Medium",
     "mama &&& yanawa *** hondai !!",
     "මම &&& යනවා *** හොඳයි !!",
     "Special Characters",
     "Symbols interspersed between valid Singlish words – verifies that symbols do not break word recognition."),

    ("Neg_0010", "Short",
     "<>|\\~`",
     "",
     "Special Characters",
     "Bracket and pipe characters – edge-case symbols that may conflict with markup or regex parsers."),

    # ── 3. VERY LONG / STRESS INPUTS (6) ─────────────────────────────────────
    ("Neg_0011", "Long",
     "mama yanawa mama yanawa mama yanawa mama yanawa mama yanawa mama yanawa mama yanawa mama yanawa mama yanawa mama yanawa",
     "මම යනවා මම යනවා මම යනවා මම යනවා මම යනවා මම යනවා මම යනවා මම යනවා මම යනවා මම යනවා",
     "Very Long Input",
     "Same short phrase repeated 10 times – stress-tests repetitive transliteration and performance."),

    ("Neg_0012", "Long",
     "api hondai api hondai api hondai api hondai api hondai api hondai api hondai api hondai api hondai api hondai api hondai api hondai api hondai api hondai api hondai",
     "අපි හොඳයි අපි හොඳයි අපි හොඳයි අපි හොඳයි අපි හොඳයි අපි හොඳයි අපි හොඳයි අපි හොඳයි අපි හොඳයි අපි හොඳයි අපි හොඳයි අපි හොඳයි අපි හොඳයි අපි හොඳයි අපි හොඳයි",
     "Very Long Input",
     "15x repeated phrase – performance and consistency check for repeated identical transliteration."),

    ("Neg_0013", "Long",
     "mama giyath naha, mama aawath naha, mama kiwwath naha, mama dunnath naha, mama karanawath naha, mama balannath naha, mama hitiyath naha",
     "මම ගියත් නැහැ, මම ආවත් නැහැ, මම කිව්වත් නැහැ, මම දුන්නත් නැහැ, මම කරනවත් නැහැ, මම බලන්නත් නැහැ, මම හිටියත් නැහැ",
     "Very Long Input",
     "Complex multi-clause Singlish sentence with repeated structure – tests multi-clause parsing accuracy."),

    ("Neg_0014", "Long",
     "koheda yanawa kohomada hitenne oya koheda innawa mama koheda yanawa api koheda yanawa oyalata koheda yanawa mekata koheda yanawa api danna naha koheda yanawa",
     "කොහේද යනවා කොහොමද හිතෙන්නේ ඔය කොහේද ඉන්නවා මම කොහේද යනවා අපි කොහේද යනවා ඔයාලාට කොහේද යනවා මේකට කොහේද යනවා අපි දන්නේ නැහැ කොහේද යනවා",
     "Very Long Input",
     "Extremely repetitive question-form input – boundary test for question-word frequency tolerance."),

    ("Neg_0015", "Long",
     "This is a very long English sentence that has absolutely no Singlish content whatsoever and should therefore produce no meaningful Sinhala transliteration output at all because every single word in it belongs to the English language",
     "This is a very long English sentence that has absolutely no Singlish content whatsoever and should therefore produce no meaningful Sinhala transliteration output at all because every single word in it belongs to the English language",
     "Very Long Input",
     "Very long purely-English paragraph – tool must not hallucinate Sinhala output for non-Singlish content."),

    ("Neg_0016", "Long",
     "a b c d e f g h i j k l m n o p q r s t u v w x y z A B C D E F G H I J K L M N O P Q R S T U V W X Y Z 0 1 2 3 4 5 6 7 8 9",
     "a b c d e f g h i j k l m n o p q r s t u v w x y z A B C D E F G H I J K L M N O P Q R S T U V W X Y Z 0 1 2 3 4 5 6 7 8 9",
     "Very Long Input",
     "All ASCII letters and digits separated by spaces – verifies the tool does not misread character sequences as Singlish."),

    # ── 4. SCRIPT INJECTION / SECURITY (6) ───────────────────────────────────
    ("Neg_0017", "Medium",
     "<script>alert('XSS')</script>",
     "<script>alert('XSS')</script>",
     "Script Injection",
     "XSS payload – tool must not execute JavaScript; should display as plain text or escape safely."),

    ("Neg_0018", "Medium",
     "'; DROP TABLE users; --",
     "'; DROP TABLE users; --",
     "Script Injection",
     "SQL injection string – must be treated as plain text; verifies no backend query manipulation."),

    ("Neg_0019", "Medium",
     "<b>mama</b> <i>yanawa</i>",
     "<b>mama</b> <i>yanawa</i>",
     "Script Injection",
     "HTML tags wrapping valid Singlish words – tool should strip or pass-through tags without translating them."),

    ("Neg_0020", "Medium",
     "{{7*7}} ${7*7} #{7*7}",
     "{{7*7}} ${7*7} #{7*7}",
     "Script Injection",
     "Template injection payloads – tests resistance to server-side template injection (SSTI) patterns."),

    ("Neg_0021", "Medium",
     "javascript:alert(1) onmouseover=alert(1)",
     "javascript:alert(1) onmouseover=alert(1)",
     "Script Injection",
     "Inline event-handler injection – verifies the tool does not interpret event attributes as executable code."),

    ("Neg_0022", "Long",
     "mama yanawa <script>document.cookie='stolen='+document.cookie</script> eka hondai",
     "මම යනවා <script>document.cookie='stolen='+document.cookie</script> ඒක හොඳයි",
     "Script Injection",
     "XSS embedded inside valid Singlish sentence – tests whether transliteration isolates and sanitizes injected code."),

    # ── 5. NON-SINGLISH / FOREIGN LANGUAGES (6) ──────────────────────────────
    ("Neg_0023", "Short",
     "مرحبا كيف حالك",
     "مرحبا كيف حالك",
     "Non-Singlish Language",
     "Arabic greeting – right-to-left script; tool should return unchanged text and not attempt transliteration."),

    ("Neg_0024", "Short",
     "நான் நன்றாக இருக்கிறேன்",
     "நான் நன்றாக இருக்கிறேன்",
     "Non-Singlish Language",
     "Tamil sentence – South Asian script; should pass through without Sinhala transliteration output."),

    ("Neg_0025", "Short",
     "你好，你怎么样？",
     "你好，你怎么样？",
     "Non-Singlish Language",
     "Mandarin Chinese – CJK characters; tool must return text unchanged since no Singlish content exists."),

    ("Neg_0026", "Medium",
     "Comment allez-vous aujourd'hui mon ami?",
     "Comment allez-vous aujourd'hui mon ami?",
     "Non-Singlish Language",
     "French sentence – Latin-script foreign language that superficially resembles English; must not transliterate."),

    ("Neg_0027", "Short",
     "Wie geht es Ihnen?",
     "Wie geht es Ihnen?",
     "Non-Singlish Language",
     "German question – non-Singlish Latin-script input; tool should leave German words unchanged."),

    ("Neg_0028", "Medium",
     "Привет как дела у тебя сегодня",
     "Привет как дела у тебя сегодня",
     "Non-Singlish Language",
     "Russian Cyrillic script – completely unrelated to Singlish; tool must handle non-Latin scripts gracefully."),

    # ── 6. GIBBERISH / RANDOM CHARACTERS (5) ─────────────────────────────────
    ("Neg_0029", "Short",
     "asdfjkl;",
     "",
     "Gibberish",
     "Random keyboard home-row characters – meaningless input; tool should return empty or unchanged."),

    ("Neg_0030", "Short",
     "qwerty uiop",
     "",
     "Gibberish",
     "QWERTY keyboard row characters – not a valid Singlish word; expect no meaningful Sinhala output."),

    ("Neg_0031", "Medium",
     "zxcvbnm asdfghjkl qwertyuiop poiuytrewq lkjhgfdsa mnbvcxz",
     "",
     "Gibberish",
     "All keyboard rows typed randomly – complete gibberish with no Singlish vocabulary; expects empty output."),

    ("Neg_0032", "Short",
     "xyzxyzxyz",
     "",
     "Gibberish",
     "Repeated nonsense character sequence – tool must not guess or fabricate a Sinhala translation."),

    ("Neg_0033", "Long",
     "fjdksla fjdksla fjdksla mama fjdksla fjdksla yanawa fjdksla fjdksla fjdksla hondai fjdksla fjdksla fjdksla",
     "fjdksla fjdksla fjdksla මම fjdksla fjdksla යනවා fjdksla fjdksla fjdksla හොඳයි fjdksla fjdksla fjdksla",
     "Gibberish",
     "Valid Singlish words scattered among gibberish tokens – tests whether tool correctly isolates known words."),

    # ── 7. MIXED SCRIPTS (5) ──────────────────────────────────────────────────
    ("Neg_0034", "Short",
     "මම yanawa",
     "මම yanawa",
     "Mixed Scripts",
     "Sinhala Unicode + Singlish Latin mix – tool input expected to be pure Singlish; mixing scripts is invalid."),

    ("Neg_0035", "Medium",
     "mama යනවා hondai හොඳයි",
     "mama යනවා හොඳයි හොඳයි",
     "Mixed Scripts",
     "Alternating Singlish and native Sinhala Unicode – tests whether tool handles already-transliterated words."),

    ("Neg_0036", "Medium",
     "api ගියා Colombo ehata 😊",
     "api ගියා Colombo එහාට 😊",
     "Mixed Scripts",
     "Native Sinhala Unicode mixed with Singlish and emoji – partially pre-translated sentence edge case."),

    ("Neg_0037", "Long",
     "mama office yanawa, ඒකට ගියා, came back, නැවත ගියා, went again, ආවා, and so on and so forth",
     "මම office යනවා, ඒකට ගියා, came back, නැවත ගියා, went again, ආවා, and so on and so forth",
     "Mixed Scripts",
     "Multi-language sentence: Singlish + Sinhala Unicode + English – complex mixed-script stress test."),

    ("Neg_0038", "Medium",
     "ك mama ك yanawa ك hondai ك",
     "ك mama ක yanawa ක හොඳයි ක",
     "Mixed Scripts",
     "Arabic single character interspersed with Singlish – RTL character mixed in LTR Singlish text."),

    # ── 8. NUMBERS AND SYMBOLS ONLY (5) ──────────────────────────────────────
    ("Neg_0039", "Short",
     "12345",
     "12345",
     "Numbers Only",
     "Pure numeric string – no Singlish words; tool should return numbers unchanged without transliteration."),

    ("Neg_0040", "Medium",
     "3.14159 2.71828 1.41421",
     "3.14159 2.71828 1.41421",
     "Numbers Only",
     "Mathematical constants as floating-point numbers – no textual Singlish content; should pass through."),

    ("Neg_0041", "Short",
     "0000000000",
     "0000000000",
     "Numbers Only",
     "Ten zeros – edge case numeric input; tool must not interpret repeated zeros as transliterable text."),

    ("Neg_0042", "Medium",
     "100 200 300 400 500 600 700 800 900 1000",
     "100 200 300 400 500 600 700 800 900 1000",
     "Numbers Only",
     "Sequence of round numbers – pure numeric list; verifies that number-only inputs are passed through intact."),

    ("Neg_0043", "Medium",
     "1+1=2 and 5*5=25 and 10/2=5",
     "1+1=2 and 5*5=25 and 10/2=5",
     "Numbers Only",
     "Math expressions with English connectors – arithmetic notation should not be transliterated."),

    # ── 9. ENCODING / UNICODE EDGE CASES (4) ─────────────────────────────────
    ("Neg_0044", "Short",
     "😂😂😂😂😂",
     "😂😂😂😂😂",
     "Encoding Edge Cases",
     "Emoji-only input (5 laughing emojis) – no text content; tool must return emojis unchanged."),

    ("Neg_0045", "Short",
     "🔥💯🎉🙌✨",
     "🔥💯🎉🙌✨",
     "Encoding Edge Cases",
     "Mixed emoji sequence only – diverse emoji set with no Singlish; expects identical pass-through output."),

    ("Neg_0046", "Medium",
     "mama yanawa 🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟",
     "මම යනවා 🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟🌟",
     "Encoding Edge Cases",
     "Singlish with 20 repeated star emojis – tests whether large emoji blocks affect transliteration logic."),

    ("Neg_0047", "Medium",
     "Z̷̢̠̹̙̳̥͝A̸͙̍L̷͍̯̮͆̔͠G̶͈͝O̷̦̕ ̶̲̏T̴͉̀̇Ë̴̤̦́X̷̨̛̥͠T̸͙͐",
     "Z̷̢̠̹̙̳̥͝A̸͙̍L̷͍̯̮͆̔͠G̶͈͝O̷̦̕ ̶̲̏T̴͉̀̇Ë̴̤̦́X̷̨̛̥͠T̸͙͐",
     "Encoding Edge Cases",
     "Zalgo/glitched unicode text – heavily composited diacritics; tool must not crash on malformed Unicode combining chars."),

    # ── 10. FORMATTING / CONTROL CHARACTERS (3) ──────────────────────────────
    ("Neg_0048", "Medium",
     "mama\nyanawa\nhondai",
     "මම\nයනවා\nහොඳයි",
     "Formatting / Control Chars",
     "Newline characters separating Singlish words – tests multi-line input handling within a single field."),

    ("Neg_0049", "Long",
     "mama yanawa\n\nmama enawa\n\nmama innawa\n\nmama karanawa\n\nmama balannawa\n\nmama kiyanawa",
     "මම යනවා\n\nමම එනවා\n\nමම ඉන්නවා\n\nමම කරනවා\n\nමම බලන්නවා\n\nමම කියනවා",
     "Formatting / Control Chars",
     "Multiple blank-line-separated Singlish sentences – paragraph structure test using double newlines."),

    ("Neg_0050", "Long",
     "mama yanawa machan, mama api giyoth oya ohoma kiyala kiwwada? naha mama eya kiwwama api hondatama giyath naha, ehema wunnath api danna naha koheda yanawa kiyala, api hondai thanama yanawa theen kiyala",
     "මම යනවා මචන්, මම අපි ගියොත් ඔය ඕහොම කියලා කිව්වද? නැහැ මම ඒය කිව්වම අපි හොඳටම ගියත් නැහැ, එහෙම වුන්නත් අපි දන්නේ නැහැ කොහේද යනවා කියලා, අපි හොඳයි තාම යනවා තීන් කියලා",
     "Formatting / Control Chars",
     "Very long complex Singlish paragraph with multiple clauses, conjunctions, and colloquial structures – ultimate stress test for transliteration accuracy and performance."),
]


def set_header_style(ws, row, num_cols):
    header_fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border


def set_data_style(ws, row, num_cols, is_odd):
    # Negative cases: light red / white alternating
    fill_color = "FDECEA" if is_odd else "FFFFFF"
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    data_font = Font(size=10, name="Calibri")
    data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = row_fill
        cell.font = data_font
        cell.alignment = data_align
        cell.border = border


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[" Test cases"]

    # Unmerge all merged cells first to avoid read-only errors
    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))

    # Clear existing data
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    # Set headers (row 1)
    headers = [
        "Test Case ID",
        "Input length type",
        "Input",
        "Expected output",
        "Actual output",
        "Status",
        "Singlish input types covered",
        "Evidence or rationale",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header

    set_header_style(ws, 1, len(headers))
    ws.row_dimensions[1].height = 35

    # Set column widths
    col_widths = [15, 18, 45, 45, 45, 10, 28, 60]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Write test cases
    for row_offset, tc in enumerate(TEST_CASES, start=2):
        tc_id, length_type, singlish, sinhala, tc_type, rationale = tc
        row_data = [tc_id, length_type, singlish, sinhala, "", "", tc_type, rationale]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_offset, column=col_idx).value = value
        set_data_style(ws, row_offset, len(headers), row_offset % 2 == 0)
        ws.row_dimensions[row_offset].height = 40

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(EXCEL_PATH)
    print(f"✅ Done! {len(TEST_CASES)} NEGATIVE test cases written to Excel.")
    print(f"   File: {EXCEL_PATH}")
    print()
    # Summary breakdown
    categories = {}
    lengths = {"Short": 0, "Medium": 0, "Long": 0}
    for tc in TEST_CASES:
        cat = tc[4]
        categories[cat] = categories.get(cat, 0) + 1
        lengths[tc[1]] = lengths.get(tc[1], 0) + 1
    print("── Category Breakdown ──")
    for cat, cnt in categories.items():
        print(f"  {cat:<35} {cnt}")
    print()
    print("── Length Breakdown ──")
    for ltype, cnt in lengths.items():
        print(f"  {ltype:<10} {cnt}")


if __name__ == "__main__":
    main()
